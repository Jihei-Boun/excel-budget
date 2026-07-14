"""예실대비표 엑셀에서 비용 항목을 읽어 파일별·통합 데이터를 만듭니다."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from schema import BudgetSchema, SummaryRule, normalize_label
from schema_detect import convert_to_number, detect_budget_schema


def find_excel_files(input_dir: Path, output_file_name: str) -> list[Path]:
    """
    입력 폴더에서 통합 대상 엑셀 파일을 찾는다.

    임시 엑셀 파일(~$로 시작)과 이미 만들어진 출력 파일은 제외한다.
    """
    excel_files = []

    for file_path in input_dir.glob("*.xlsx"):
        if file_path.name.startswith("~$"):
            continue

        if file_path.name == output_file_name:
            continue

        excel_files.append(file_path)

    return sorted(excel_files)


def extract_project_name(file_path: Path) -> str:
    """
    파일명에서 과제명을 추출한다.

    예:
    03_트윈_예실대비표.xlsx -> 트윈
    04_분산_예실대비표.xlsx -> 분산
    """
    file_name = file_path.stem
    parts = file_name.split("_")

    if len(parts) >= 3:
        return "_".join(parts[1:-1])

    return file_name


def _empty_values(column_count: int) -> list[int]:
    return [0] * column_count


def _add_values(left: list[int | float], right: list[int | float]) -> list[int | float]:
    return [a + b for a, b in zip(left, right, strict=True)]


def _schemas_compatible(left: BudgetSchema, right: BudgetSchema) -> bool:
    return left.amount_headers == right.amount_headers


def read_budget_file(
    file_path: Path,
    schema: BudgetSchema | None = None,
) -> dict[str, Any]:
    """
    엑셀 파일 한 개에서 비용 코드 기준 상세 행을 읽는다.

    schema 가 없으면 LLM/규칙 기반으로 자동 탐지한다.
    """
    workbook = load_workbook(file_path, data_only=True)

    try:
        worksheet = workbook.active
        detected = schema or detect_budget_schema(worksheet)

        detail_rows: list[dict[str, Any]] = []
        current_category = ""
        max_row = worksheet.max_row or 200
        start_row = detected.header_row_count + 1

        for row_number in range(start_row, max_row + 1):
            col_category = worksheet.cell(
                row=row_number,
                column=detected.category_col,
            ).value
            col_code = worksheet.cell(
                row=row_number,
                column=detected.code_col,
            ).value
            col_name = worksheet.cell(
                row=row_number,
                column=detected.name_col,
            ).value

            label = normalize_label(col_category)
            code_text = "" if col_code is None else str(col_code).strip()

            if label in detected.summary_normalized or label in detected.subtotal_normalized:
                continue

            if code_text == "":
                if col_category is not None and str(col_category).strip() != "":
                    current_category = str(col_category).strip()
                continue

            if col_category is not None and str(col_category).strip() != "":
                current_category = str(col_category).strip()

            values = [
                convert_to_number(
                    worksheet.cell(row=row_number, column=column_number).value
                )
                for column_number in detected.amount_col_numbers
            ]

            detail_rows.append(
                {
                    "비목분류": current_category,
                    "비용코드": code_text,
                    "비용명": "" if col_name is None else str(col_name).strip(),
                    "값": values,
                }
            )

        if not detail_rows:
            raise ValueError("비용 코드가 있는 상세 행을 찾지 못했습니다.")

        return {
            "과제명": extract_project_name(file_path),
            "원본파일": file_path.name,
            "headers": detected.amount_headers,
            "detail_rows": detail_rows,
            "schema": detected,
        }

    finally:
        workbook.close()


def _sum_by_codes(
    detail_rows: list[dict[str, Any]],
    codes: set[str],
    column_count: int,
) -> list[int | float]:
    total = _empty_values(column_count)
    for row in detail_rows:
        if row["비용코드"] in codes:
            total = _add_values(total, row["값"])
    return total


def _apply_summary_rule(
    rule: SummaryRule,
    detail_rows: list[dict[str, Any]],
    column_count: int,
    code_summary_totals: list[list[int | float]],
) -> list[int | float]:
    if rule.composition == "all":
        total = _empty_values(column_count)
        for row in detail_rows:
            total = _add_values(total, row["값"])
        return total

    if rule.composition == "codes":
        return _sum_by_codes(detail_rows, set(rule.codes), column_count)

    # remainder: 전체 - 지금까지의 codes 요약 합
    grand_total = _empty_values(column_count)
    for row in detail_rows:
        grand_total = _add_values(grand_total, row["값"])

    absorbed = _empty_values(column_count)
    for values in code_summary_totals:
        absorbed = _add_values(absorbed, values)

    return [
        total - part
        for total, part in zip(grand_total, absorbed, strict=True)
    ]


def build_table_rows(
    detail_rows: list[dict[str, Any]],
    schema: BudgetSchema,
) -> list[dict[str, Any]]:
    """상세 행으로부터 소계·요약 행을 포함한 표 행을 만든다."""
    if not detail_rows:
        return []

    column_count = len(detail_rows[0]["값"])
    table_rows: list[dict[str, Any]] = []

    category_order: list[str] = []
    rows_by_category: dict[str, list[dict[str, Any]]] = {}

    for row in detail_rows:
        category = row["비목분류"] or "기타"

        if category not in rows_by_category:
            category_order.append(category)
            rows_by_category[category] = []

        rows_by_category[category].append(row)

    subtotal_label = schema.subtotal_labels[0] if schema.subtotal_labels else "소계"

    for category in category_order:
        category_rows = sorted(
            rows_by_category[category],
            key=lambda item: (
                int(item["비용코드"])
                if str(item["비용코드"]).isdigit()
                else str(item["비용코드"])
            ),
        )
        category_total = _empty_values(column_count)
        first_in_category = True

        for row in category_rows:
            values = list(row["값"])
            category_total = _add_values(category_total, values)

            table_rows.append(
                {
                    "행종류": "detail",
                    "비목분류": category if first_in_category else "",
                    "비용코드": row["비용코드"],
                    "비용명": row["비용명"],
                    "값": values,
                }
            )
            first_in_category = False

        table_rows.append(
            {
                "행종류": "subtotal",
                "비목분류": subtotal_label,
                "비용코드": "",
                "비용명": "",
                "값": category_total,
            }
        )

    code_summary_totals: list[list[int | float]] = []

    for rule in schema.summary_rules:
        values = _apply_summary_rule(
            rule,
            detail_rows,
            column_count,
            code_summary_totals,
        )

        if rule.composition == "codes":
            code_summary_totals.append(values)

        table_rows.append(
            {
                "행종류": "summary",
                "비목분류": rule.label,
                "비용코드": "",
                "비용명": "",
                "값": values,
            }
        )

    return table_rows


def merge_by_cost_code(file_tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    여러 파일의 상세 행을 비용 코드 기준으로 합산한다.

    한쪽 파일에만 있는 항목은 없는 쪽을 0으로 본다.
    """
    merged: dict[str, dict[str, Any]] = {}
    code_order: list[str] = []

    for file_table in file_tables:
        for row in file_table["detail_rows"]:
            code = row["비용코드"]

            if code not in merged:
                code_order.append(code)
                merged[code] = {
                    "비목분류": row["비목분류"],
                    "비용코드": code,
                    "비용명": row["비용명"],
                    "값": list(row["값"]),
                }
            else:
                merged[code]["값"] = _add_values(merged[code]["값"], row["값"])

                if not merged[code]["비목분류"] and row["비목분류"]:
                    merged[code]["비목분류"] = row["비목분류"]

                if not merged[code]["비용명"] and row["비용명"]:
                    merged[code]["비용명"] = row["비용명"]

    return [merged[code] for code in code_order]


def collect_all_files(
    excel_files: list[Path],
) -> tuple[list[str], list[dict[str, Any]], list[dict[str, Any]], BudgetSchema | None]:
    """
    여러 엑셀 파일을 읽어 원본 표들과 통합 표를 만든다.

    Returns:
        (금액 헤더, 원본 표 목록, 통합 표 행, 기준 스키마)
    """
    file_tables: list[dict[str, Any]] = []
    base_schema: BudgetSchema | None = None
    errors: list[str] = []
    used_sheet_names: set[str] = set()

    for file_path in excel_files:
        try:
            # 첫 성공 파일에서만 스키마를 탐지하고, 이후 파일은 재사용한다.
            file_table = read_budget_file(file_path, schema=base_schema)
            schema: BudgetSchema = file_table["schema"]

            if base_schema is None:
                base_schema = schema
            elif not _schemas_compatible(base_schema, schema):
                raise ValueError(
                    "헤더 구조가 기준 파일과 다릅니다. "
                    f"기준={base_schema.amount_headers}, "
                    f"실제={schema.amount_headers}"
                )

            sheet_name = file_table["과제명"]
            if sheet_name in used_sheet_names:
                sheet_name = f"{sheet_name}_{len(used_sheet_names) + 1}"

            used_sheet_names.add(sheet_name)

            file_tables.append(
                {
                    **file_table,
                    "시트명": sheet_name,
                    "table_rows": build_table_rows(
                        file_table["detail_rows"],
                        base_schema,
                    ),
                }
            )
            print(f"[성공] {file_path.name}")

        except Exception as error:
            errors.append(f"{file_path.name}: {error}")
            print(f"[실패] {file_path.name}: {error}")

    if errors:
        print("\n처리하지 못한 파일")
        for error_message in errors:
            print(f"- {error_message}")

    if base_schema is None:
        return [], [], [], None

    merged_details = merge_by_cost_code(file_tables)
    integrated_rows = build_table_rows(merged_details, base_schema)

    return base_schema.amount_headers, file_tables, integrated_rows, base_schema
