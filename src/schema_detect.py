"""엑셀 표 구조를 LLM/추론으로 자동 탐지합니다."""

from __future__ import annotations

import itertools
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from llm_client import chat_json, llm_enabled
from schema import BudgetSchema, SummaryRule, display_label, normalize_label


def convert_to_number(value: Any) -> int | float:
    """엑셀 셀 값을 숫자로 변환한다."""
    if value is None:
        return 0

    if isinstance(value, (int, float)):
        return value

    text = str(value).strip().replace(",", "")
    if text == "":
        return 0

    try:
        number = float(text)
        if number.is_integer():
            return int(number)
        return number
    except ValueError as error:
        raise ValueError(f"숫자로 변환할 수 없는 값입니다: {value!r}") from error


def worksheet_preview(worksheet: Worksheet, max_rows: int = 40) -> list[list[Any]]:
    """LLM/추론용으로 시트를 2차원 리스트로 요약한다."""
    max_row = min(worksheet.max_row or 0, max_rows)
    max_col = worksheet.max_column or 0
    rows: list[list[Any]] = []

    for row_number in range(1, max_row + 1):
        row = [
            worksheet.cell(row=row_number, column=column_number).value
            for column_number in range(1, max_col + 1)
        ]
        rows.append(row)

    return rows


def build_amount_headers(
    worksheet: Worksheet,
    header_row_count: int,
    amount_start_col: int,
) -> tuple[list[int], list[str]]:
    """헤더 행과 금액 시작 열을 기준으로 금액 열 헤더를 만든다."""
    max_column = worksheet.max_column or 0
    parent_name = ""
    col_numbers: list[int] = []
    headers: list[str] = []

    top_row = 1
    bottom_row = 2 if header_row_count >= 2 else 1

    for column_number in range(amount_start_col, max_column + 1):
        top = worksheet.cell(row=top_row, column=column_number).value
        bottom = (
            worksheet.cell(row=bottom_row, column=column_number).value
            if header_row_count >= 2
            else None
        )

        top_text = str(top).strip() if top is not None else ""
        bottom_text = str(bottom).strip() if bottom is not None else ""

        if top_text:
            parent_name = top_text

        if not top_text and not bottom_text:
            break

        if bottom_text and parent_name and top_text == "":
            header_name = f"{parent_name}_{bottom_text}"
        elif bottom_text and top_text:
            header_name = f"{top_text}_{bottom_text}"
        elif bottom_text:
            header_name = bottom_text
        else:
            header_name = top_text

        col_numbers.append(column_number)
        headers.append(header_name)

    if not headers:
        raise ValueError("금액 열 헤더를 만들지 못했습니다.")

    return col_numbers, headers


def _looks_like_code(value: Any) -> bool:
    """비용 코드처럼 보이는 짧은 값인지 판별한다."""
    if value is None:
        return False

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    if isinstance(value, int):
        return 0 < abs(value) < 10000

    text = str(value).strip()
    if not text or len(text) > 8:
        return False

    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]

    compact = text.replace("-", "").replace("_", "")

    # 한글/공백 라벨은 코드가 아니다. ASCII 숫자·영문만 허용.
    if not compact.isascii() or not compact.isalnum():
        return False

    if compact.isdigit():
        return 0 < len(compact) <= 4

    return True


def _is_numeric_cell(value: Any) -> bool:
    if value is None or value == "":
        return False
    try:
        convert_to_number(value)
        return True
    except ValueError:
        return False


def infer_layout(preview: list[list[Any]]) -> dict[str, Any]:
    """헤더/열 위치를 규칙 기반으로 추정한다."""
    if len(preview) < 3:
        raise ValueError("표가 너무 짧아 구조를 추론할 수 없습니다.")

    header_row_count = 1
    if len(preview) >= 2:
        second = preview[1]
        # 2행에 부모 없이 자식 헤더만 있으면 2단 헤더로 본다.
        non_empty = [cell for cell in second if cell not in (None, "")]
        if non_empty and all(
            not _is_numeric_cell(cell) and not _looks_like_code(cell)
            for cell in non_empty
        ):
            header_row_count = 2

    data_rows = preview[header_row_count:]
    max_col = max((len(row) for row in preview), default=0)

    code_scores = []
    for col_idx in range(max_col):
        code_like = 0
        numeric = 0
        for row in data_rows:
            if col_idx >= len(row):
                continue
            value = row[col_idx]
            if _looks_like_code(value) and not _is_numeric_cell(value):
                # 숫자처럼 보이는 짧은 코드(121)도 코드로 취급
                code_like += 1
            elif _looks_like_code(value):
                code_like += 1
            if _is_numeric_cell(value):
                numeric += 1
        code_scores.append((code_like, -numeric, col_idx))

    # 비용코드: 코드 형태가 많고, 금액 열보다는 왼쪽
    code_col_idx = max(code_scores)[2] if code_scores else 1
    category_col_idx = 0 if code_col_idx > 0 else 0
    name_col_idx = code_col_idx + 1 if code_col_idx + 1 < max_col else code_col_idx

    amount_start_idx = None
    for col_idx in range(code_col_idx + 1, max_col):
        numeric_count = sum(
            1
            for row in data_rows
            if col_idx < len(row) and _is_numeric_cell(row[col_idx])
        )
        if numeric_count >= max(3, len(data_rows) // 4):
            amount_start_idx = col_idx
            break

    if amount_start_idx is None:
        amount_start_idx = max(code_col_idx + 2, 3)

    # name 열이 금액 시작과 겹치면 보정
    if name_col_idx >= amount_start_idx:
        name_col_idx = max(code_col_idx + 1, amount_start_idx - 1)

    return {
        "header_row_count": header_row_count,
        "category_col": category_col_idx + 1,
        "code_col": code_col_idx + 1,
        "name_col": name_col_idx + 1,
        "amount_start_col": amount_start_idx + 1,
    }


def collect_raw_rows(
    worksheet: Worksheet,
    layout: dict[str, Any],
    amount_col_numbers: list[int],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """상세/소계/요약 후보 행을 분리한다."""
    category_col = layout["category_col"]
    code_col = layout["code_col"]
    name_col = layout["name_col"]
    start_row = layout["header_row_count"] + 1
    max_row = worksheet.max_row or 0

    details: list[dict[str, Any]] = []
    subtotals: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    current_category = ""

    for row_number in range(start_row, max_row + 1):
        label = worksheet.cell(row=row_number, column=category_col).value
        code = worksheet.cell(row=row_number, column=code_col).value
        name = worksheet.cell(row=row_number, column=name_col).value
        values = [
            convert_to_number(worksheet.cell(row=row_number, column=column).value)
            for column in amount_col_numbers
        ]

        label_text = "" if label is None else str(label).strip()
        code_text = "" if code is None else str(code).strip()
        name_text = "" if name is None else str(name).strip()
        normalized = normalize_label(label_text)
        has_amount = any(value != 0 for value in values)

        if code_text:
            if label_text:
                current_category = label_text
            details.append(
                {
                    "row": row_number,
                    "비목분류": current_category,
                    "비용코드": code_text,
                    "비용명": name_text,
                    "값": values,
                }
            )
            continue

        if not label_text:
            continue

        if "소계" in normalized:
            subtotals.append(
                {
                    "row": row_number,
                    "label": label_text,
                    "값": values,
                }
            )
            continue

        if has_amount:
            summaries.append(
                {
                    "row": row_number,
                    "label": label_text,
                    "값": values,
                }
            )
            if label_text and "소계" not in normalized:
                # 비목분류만 있는 행은 상세로 안 가므로 카테고리만 갱신하지 않음
                pass
            continue

        # 금액 없는 분류 헤더
        current_category = label_text

    return details, subtotals, summaries


def _values_close(left: list[int | float], right: list[int | float], tol: float = 1.0) -> bool:
    if len(left) != len(right):
        return False
    return all(abs(a - b) <= tol for a, b in zip(left, right, strict=True))


def _find_code_subset(
    details: list[dict[str, Any]],
    target: list[int | float],
) -> list[str] | None:
    """요약 금액과 일치하는 비용 코드 부분 집합을 찾는다."""
    n = len(details)
    if n == 0:
        return None

    # 코드 수가 많으면 첫 금액 열 기준으로 가지치기한 뒤 검증
    if n <= 20:
        for size in range(1, n + 1):
            for combo in itertools.combinations(range(n), size):
                total = [0] * len(target)
                for idx in combo:
                    for col, value in enumerate(details[idx]["값"]):
                        total[col] += value
                if _values_close(total, target):
                    return [details[idx]["비용코드"] for idx in combo]
        return None

    # 큰 표: 단일 열 숫자 DP 후 전체 열 검증
    primary = [int(round(float(row["값"][0]))) for row in details]
    goal = int(round(float(target[0])))
    dp: dict[int, tuple[int, ...]] = {0: tuple()}

    for index, amount in enumerate(primary):
        updates: dict[int, tuple[int, ...]] = {}
        for current, chosen in dp.items():
            nxt = current + amount
            if nxt > goal:
                continue
            if nxt not in dp and nxt not in updates:
                updates[nxt] = chosen + (index,)
        dp.update(updates)
        if goal in dp:
            break

    if goal not in dp:
        return None

    indices = dp[goal]
    total = [0] * len(target)
    for idx in indices:
        for col, value in enumerate(details[idx]["값"]):
            total[col] += value

    if not _values_close(total, target):
        return None

    return [details[idx]["비용코드"] for idx in indices]


def infer_summary_rules(
    details: list[dict[str, Any]],
    summaries: list[dict[str, Any]],
) -> list[SummaryRule]:
    """하단 요약 행의 구성 규칙을 금액 일치로 추론한다."""
    if not summaries:
        return [
            SummaryRule(label="합계", composition="all"),
        ]

    all_total = [0] * len(details[0]["값"])
    for row in details:
        for idx, value in enumerate(row["값"]):
            all_total[idx] += value

    rules: list[SummaryRule] = []

    for item in summaries:
        label = display_label(item["label"])
        values = item["값"]

        if _values_close(values, all_total):
            rules.append(SummaryRule(label=label, composition="all"))
            continue

        # 이미 찾은 codes 요약의 나머지를 먼저 검사한다.
        # (외부유출액처럼 코드 목록을 하드코딩하지 않기 위함)
        known_sum = [0] * len(all_total)
        for rule in rules:
            if rule.composition != "codes":
                continue
            for row in details:
                if row["비용코드"] in rule.codes:
                    for idx, value in enumerate(row["값"]):
                        known_sum[idx] += value

        remainder = [
            total - known for total, known in zip(all_total, known_sum, strict=True)
        ]
        if any(rule.composition == "codes" for rule in rules) and _values_close(
            values,
            remainder,
        ):
            rules.append(SummaryRule(label=label, composition="remainder"))
            continue

        codes = _find_code_subset(details, values)
        if codes:
            rules.append(
                SummaryRule(label=label, composition="codes", codes=codes)
            )
            continue

        rules.append(SummaryRule(label=label, composition="remainder"))

    if not any(rule.composition == "all" for rule in rules):
        rules.append(SummaryRule(label="합계", composition="all"))

    return [
        SummaryRule(
            label=display_label(rule.label) or rule.label,
            composition=rule.composition,
            codes=rule.codes,
        )
        for rule in rules
    ]


def detect_schema_heuristic(worksheet: Worksheet) -> BudgetSchema:
    """LLM 없이 표 구조를 추론한다."""
    preview = worksheet_preview(worksheet)
    layout = infer_layout(preview)
    amount_cols, headers = build_amount_headers(
        worksheet,
        layout["header_row_count"],
        layout["amount_start_col"],
    )
    details, subtotals, summaries = collect_raw_rows(
        worksheet,
        layout,
        amount_cols,
    )

    if not details:
        raise ValueError("상세 행을 찾지 못했습니다.")

    subtotal_labels = sorted(
        {display_label(item["label"]) for item in subtotals} or {"소계"}
    )
    summary_rules = infer_summary_rules(details, summaries)

    return BudgetSchema(
        header_row_count=layout["header_row_count"],
        category_col=layout["category_col"],
        code_col=layout["code_col"],
        name_col=layout["name_col"],
        amount_start_col=layout["amount_start_col"],
        amount_headers=headers,
        amount_col_numbers=amount_cols,
        subtotal_labels=subtotal_labels,
        summary_rules=summary_rules,
        source="heuristic",
    )


def _schema_from_llm_payload(
    payload: dict[str, Any],
    worksheet: Worksheet,
    details: list[dict[str, Any]],
    summaries: list[dict[str, Any]],
    subtotals: list[dict[str, Any]],
) -> BudgetSchema:
    header_row_count = int(payload.get("header_row_count", 2))
    category_col = int(payload["category_col"])
    code_col = int(payload["code_col"])
    name_col = int(payload["name_col"])
    amount_start_col = int(payload["amount_start_col"])

    amount_cols, headers = build_amount_headers(
        worksheet,
        header_row_count,
        amount_start_col,
    )

    raw_rules = payload.get("summary_rules") or []
    summary_rules: list[SummaryRule] = []

    for item in raw_rules:
        composition = item.get("composition", "remainder")
        if composition not in {"codes", "remainder", "all"}:
            composition = "remainder"
        summary_rules.append(
            SummaryRule(
                label=display_label(item.get("label", "")),
                composition=composition,
                codes=[str(code).strip() for code in item.get("codes", [])],
            )
        )

    # LLM 요약 규칙은 금액으로 재검증·보정한다.
    verified = infer_summary_rules(details, summaries)
    if verified:
        # 라벨은 LLM/원본 요약을 우선, 구성은 검증 결과를 우선
        label_by_index = {
            idx: rule.label
            for idx, rule in enumerate(summary_rules)
            if rule.label
        }
        summary_rules = []
        for idx, rule in enumerate(verified):
            label = label_by_index.get(idx, rule.label)
            if idx < len(summaries):
                label = summaries[idx]["label"]
            summary_rules.append(
                SummaryRule(
                    label=display_label(label),
                    composition=rule.composition,
                    codes=rule.codes,
                )
            )

    subtotal_labels = [
        display_label(label)
        for label in payload.get("subtotal_labels", [])
        if display_label(label)
    ]
    if not subtotal_labels:
        subtotal_labels = sorted(
            {display_label(item["label"]) for item in subtotals} or {"소계"}
        )

    return BudgetSchema(
        header_row_count=header_row_count,
        category_col=category_col,
        code_col=code_col,
        name_col=name_col,
        amount_start_col=amount_start_col,
        amount_headers=headers,
        amount_col_numbers=amount_cols,
        subtotal_labels=subtotal_labels,
        summary_rules=summary_rules,
        source="llm",
    )


def detect_schema_with_llm(worksheet: Worksheet) -> BudgetSchema:
    """LLM으로 표 구조를 읽고, 금액으로 검증한 스키마를 반환한다."""
    preview = worksheet_preview(worksheet)
    fallback_layout = infer_layout(preview)
    amount_cols, _ = build_amount_headers(
        worksheet,
        fallback_layout["header_row_count"],
        fallback_layout["amount_start_col"],
    )
    details, subtotals, summaries = collect_raw_rows(
        worksheet,
        fallback_layout,
        amount_cols,
    )

    preview_text_rows = []
    for row_idx, row in enumerate(preview, start=1):
        cells = []
        for col_idx, value in enumerate(row, start=1):
            if value is None or value == "":
                continue
            cells.append(f"{col_idx}:{value}")
        preview_text_rows.append(f"R{row_idx} | " + " | ".join(cells))

    summary_hint = [
        {
            "label": item["label"],
            "first_amount": item["값"][0] if item["값"] else 0,
        }
        for item in summaries
    ]
    detail_hint = [
        {
            "code": item["비용코드"],
            "name": item["비용명"],
            "category": item["비목분류"],
            "first_amount": item["값"][0] if item["값"] else 0,
        }
        for item in details
    ]

    system = (
        "당신은 예실대비표 엑셀 구조를 분석하는 도우미입니다. "
        "반드시 JSON 객체만 반환하세요."
    )
    prompt = f"""
다음 예실대비표 미리보기를 보고 구조를 JSON으로 알려주세요.

요구 스키마:
{{
  "header_row_count": 2,
  "category_col": 1,
  "code_col": 2,
  "name_col": 3,
  "amount_start_col": 4,
  "subtotal_labels": ["소 계"],
  "summary_rules": [
    {{"label": "요약행라벨1", "composition": "codes", "codes": ["코드A", "코드B"]}},
    {{"label": "요약행라벨2", "composition": "remainder", "codes": []}},
    {{"label": "요약행라벨3", "composition": "all", "codes": []}}
  ]
}}

규칙:
- composition 은 codes | remainder | all 중 하나
- codes: 해당 요약 금액이 특정 비용코드 합인 경우 (상세 힌트의 금액으로 판단)
- all: 모든 상세 행 합
- remainder: 전체에서 다른 codes 요약을 뺀 나머지
- 열 번호는 1부터 시작
- 파일에 실제로 보이는 라벨/코드만 사용

미리보기:
{chr(10).join(preview_text_rows)}

상세 행 힌트:
{detail_hint}

요약 행 힌트:
{summary_hint}

소계 행 라벨:
{[item["label"] for item in subtotals]}
""".strip()

    payload = chat_json(prompt, system=system)
    return _schema_from_llm_payload(
        payload,
        worksheet,
        details,
        summaries,
        subtotals,
    )


def detect_budget_schema(worksheet: Worksheet) -> BudgetSchema:
    """
    표 구조를 자동 탐지한다.

    1) LLM 사용 가능하면 LLM 분석 + 금액 검증
    2) 실패 시 규칙 기반 추론으로 대체
    """
    if llm_enabled():
        try:
            schema = detect_schema_with_llm(worksheet)
            print(
                "[스키마] LLM 탐지 성공: "
                f"code_col={schema.code_col}, "
                f"amount_start={schema.amount_start_col}, "
                f"요약규칙={[(r.label, r.composition, r.codes) for r in schema.summary_rules]}"
            )
            return schema
        except Exception as error:
            print(f"[스키마] LLM 탐지 실패, 규칙 기반으로 전환: {error}")

    schema = detect_schema_heuristic(worksheet)
    print(
        "[스키마] 규칙 기반 탐지: "
        f"code_col={schema.code_col}, "
        f"amount_start={schema.amount_start_col}, "
        f"요약규칙={[(r.label, r.composition, r.codes) for r in schema.summary_rules]}"
    )
    return schema
