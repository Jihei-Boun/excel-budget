"""통합 결과 엑셀을 작성하고 저장합니다."""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_budget_sheet(
    workbook: Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    """비용 항목 표를 엑셀 시트 한 개에 작성한다."""
    worksheet = workbook.create_sheet(title=sheet_name)

    worksheet.append(
        [
            "비목분류",
            "비용코드",
            "비용명",
            *headers,
        ]
    )

    for item in rows:
        worksheet.append(
            [
                item["비목분류"],
                item["비용코드"],
                item["비용명"],
                *item["값"],
            ]
        )

    format_worksheet(worksheet, rows)


def format_worksheet(worksheet, rows: list[dict[str, Any]]) -> None:
    """결과 시트의 기본 서식을 적용한다."""
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )
    subtotal_fill = PatternFill(
        fill_type="solid",
        fgColor="F2F2F2",
    )
    summary_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "D2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for row_offset, item in enumerate(rows, start=2):
        row_kind = item.get("행종류", "detail")

        if row_kind in {"subtotal", "summary"}:
            fill = summary_fill if row_kind == "summary" else subtotal_fill

            for cell in worksheet[row_offset]:
                cell.font = Font(bold=True)
                cell.fill = fill

    for row in worksheet.iter_rows(
        min_row=2,
        min_col=4,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            cell.number_format = "#,##0;[Red]-#,##0"

    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 10
    worksheet.column_dimensions["C"].width = 22

    for column_number in range(4, worksheet.max_column + 1):
        column_letter = get_column_letter(column_number)
        worksheet.column_dimensions[column_letter].width = 16

    worksheet.row_dimensions[1].height = 35


def save_integrated_workbook(
    headers: list[str],
    file_tables: list[dict[str, Any]],
    integrated_rows: list[dict[str, Any]],
    output_file: Path,
) -> None:
    """원본 표 N개와 통합 표 1개를 각각 시트로 저장한다."""
    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for file_table in file_tables:
        write_budget_sheet(
            workbook,
            file_table["시트명"],
            headers,
            file_table["table_rows"],
        )

    write_budget_sheet(
        workbook,
        "통합",
        headers,
        integrated_rows,
    )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
